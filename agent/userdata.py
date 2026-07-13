"""Bring-your-own-data: register an uploaded table so the SAME agent can query + model it.

Given a pandas DataFrame we (1) sanitize identifiers, (2) materialize it as a table in a fresh,
session-scoped DuckDB file, and (3) build a semantic-catalog dict of the same shape the RAG layer
expects. `run_analysis(question, catalog=..., db_path=...)` then runs the full pipeline — SQL,
guardrail, and every model (A/B, non-inferiority, regression, survival, …) — over the user's data.

The catalog carries no PHI to disk; only column names + a few example values reach the LLM at
query time, so the UI must warn users to upload non-sensitive data only on the public deploy.
"""
from __future__ import annotations

import re
import tempfile
import unicodedata
from pathlib import Path

import duckdb
import pandas as pd

MAX_ROWS = 200_000
MAX_COLS = 60
MAX_SHEET_CELLS = 10_000_000     # declared Excel-sheet size gate — parsed BEFORE any cell is read


def read_upload(file, name: str) -> tuple[pd.DataFrame, list[str]]:
    """Parse an uploaded CSV/Excel with bounded work, returning (df, user-facing notes).

    The UI's 50 MB cap measures the file as uploaded — for an xlsx that is the COMPRESSED size, and
    a small crafted workbook (highly repetitive cells) can inflate to gigabytes when parsed. So for
    Excel the sheet's declared dimensions are checked read-only before parsing, and both formats
    read at most MAX_ROWS + 1 rows (the +1 keeps truncation detectable for the caller's notice)."""
    notes: list[str] = []
    if str(name).lower().endswith(".csv"):
        return pd.read_csv(file, nrows=MAX_ROWS + 1), notes
    import openpyxl
    wb = openpyxl.load_workbook(file, read_only=True)
    try:
        ws = wb.worksheets[0]
        if len(wb.sheetnames) > 1:
            notes.append(f"This workbook has {len(wb.sheetnames)} sheets — only the first "
                         f"(**{wb.sheetnames[0]}**) was analyzed.")
        cells = (ws.max_row or 0) * (ws.max_column or 0)
        if cells > MAX_SHEET_CELLS:
            raise ValueError(
                f"the first sheet declares {ws.max_row:,} rows × {ws.max_column:,} columns "
                f"(~{cells:,} cells), over the {MAX_SHEET_CELLS:,}-cell limit — pre-aggregate or "
                "export a smaller CSV instead.")
    finally:
        wb.close()
    file.seek(0)
    return pd.read_excel(file, sheet_name=0, nrows=MAX_ROWS + 1), notes


_RESERVED = {
    "select", "from", "where", "table", "order", "group", "by", "join", "on", "and", "or", "not",
    "insert", "update", "delete", "drop", "create", "alter", "index", "view", "as", "in", "is", "null",
    "case", "when", "then", "else", "end", "union", "all", "distinct", "having", "limit", "offset",
    "left", "right", "inner", "outer", "full", "cross", "using", "values", "set", "into", "column",
    "primary", "key", "foreign", "references", "default", "check", "unique", "desc", "asc", "between",
    "like", "exists", "any", "some", "cast", "with", "over", "partition", "pivot", "sample", "row",
}


def _sanitize(name: str, fallback: str = "col") -> str:
    # Transliterate accented latin to its ascii base (Âge→age) before the ascii-only pass; pure
    # non-latin scripts (日本語) collapse to empty and fall back to `fallback`/dedup below.
    ascii_name = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^0-9a-zA-Z]+", "_", ascii_name).strip("_").lower()
    if not s:
        s = fallback
    if s[0].isdigit():
        s = f"c_{s}"
    if s in _RESERVED:              # a keyword col name would break the agent's unquoted `SELECT <col>`
        s = f"{s}_"                 # group→group_ ; dedup in sanitize_columns still guarantees uniqueness
    return s[:40]


def sanitize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns to safe, UNIQUE snake_case identifiers so generated SQL needs no quoting."""
    df = df.iloc[:, :MAX_COLS].copy()
    used: set[str] = set()
    out = []
    for c in df.columns:
        base = _sanitize(c)
        name, i = base, 1
        while name in used:                        # guarantee uniqueness even against suffixed names
            name = f"{base}_{i}"
            i += 1
        used.add(name)
        out.append(name)
    df.columns = out
    return df


def build_catalog(df: pd.DataFrame, table: str) -> dict:
    """A one-table catalog dict shaped exactly like the RAG layer's semantic_catalog.json."""
    columns = []
    for c in df.columns:
        s = df[c]
        examples = [str(v)[:28] for v in pd.unique(s.dropna())[:3]]
        columns.append({"name": c, "type": str(s.dtype), "description": "", "example_values": examples})
    table_doc = {
        "name": table,
        "description": f"User-uploaded dataset — {len(df):,} rows, {len(df.columns)} columns.",
        "primary_key": [],
        "foreign_keys": [],
        "columns": columns,
    }
    return {"tables": [table_doc], "metrics": []}


def prepare_upload(df: pd.DataFrame, filename: str = "user_data") -> tuple[Path, dict, str, pd.DataFrame]:
    """Materialize the DataFrame as a DuckDB table and return (db_path, catalog, table_name, cleaned_df)."""
    if df is None or df.shape[1] == 0:
        raise ValueError("The uploaded file has no columns to analyze.")
    df = sanitize_columns(df).head(MAX_ROWS)
    table = _sanitize(Path(str(filename)).stem, fallback="user_data")
    if table in _RESERVED:                          # a keyword table name would break generated `FROM {table}`
        table = f"t_{table}"
    db_path = Path(tempfile.mkdtemp(prefix="byod_")) / "data.duckdb"
    con = duckdb.connect(str(db_path))
    try:
        con.register("_uploaded", df)
        con.execute(f'CREATE TABLE "{table}" AS SELECT * FROM _uploaded')   # quote → keyword-safe DDL
        con.unregister("_uploaded")
    finally:
        con.close()
    return db_path, build_catalog(df, table), table, df
